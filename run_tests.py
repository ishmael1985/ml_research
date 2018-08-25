import os, re, sys, argparse

from openpyxl import load_workbook
from shutil import move, copy2, rmtree
from collections import OrderedDict
import io
import json

os.chdir(sys.path[0])   # Ensure script is executed in its directory
sys.path.append('extras')

import perf_analysis

parser = argparse.ArgumentParser(description='Test automation for SR')
parser.add_argument('--worksheet',
                    type=str,
                    required=True,
                    help="MS Excel spreadsheet containing tests")
parser.add_argument('--training_images',
                    type=str,
                    required=False,
                    help="path to training dataset")
parser.add_argument('--train_csv',
                    type=str,
                    required=False,
                    help="csv file containing specified set of images to train")
parser.add_argument('--test_images',
                    type=str,
                    required=False,
                    help="path to test dataset")
parser.add_argument('--test_csv',
                    type=str,
                    required=False,
                    help="csv file containing specified set of images to test")
parser.add_argument('--test_ids',
                    type=str,
                    required=False,
                    nargs='+',
                    help="specify only certain test IDs to run")
parser.add_argument('--analyze_only',
                    action='store_true',
                    help="run in analysis mode only")
parser.add_argument('--plot',
                    action='store_true',
                    help="plot graphs during analysis mode")
parser.add_argument('--reuse_model',
                    action='store_true',
                    help="reuse models trained with non-augmented datasets")
parser.add_argument('--reuse_test',
                    action='store_true',
                    help="reuse tests")

opt = parser.parse_args()
wb = load_workbook(filename=opt.worksheet)
ws = wb.active
test_ids = OrderedDict()
saved_models = {}
saved_tests = {}
scale_factors = []

def analyze_results(id):
    real_stdout = sys.stdout
    fake_stdout = io.StringIO()

    # If a list of scale factors has been specified, record only the last one
    try:
        sys.stdout = fake_stdout
        args = ['--results_folder', 'results/{}'.format(id),
                '--scale', scale_factors[-1]]
        if opt.plot:
            os.makedirs('results/{}/graphs'.format(id), exist_ok=True)
            args = args + ['--save_folder', 'results/{}/graphs'.format(id)]
            
        perf_analysis.main(args)
    finally:
        sys.stdout = real_stdout
        output = fake_stdout.getvalue()
        fake_stdout.close()

        summary_regex = re.compile(
            'Global average performance difference :\s+'
            '(?P<psnr_diff>\d+\.\d+)'
            '.*?'
            'Estimated total standard deviation :\s+'
            '(?P<deviation>\d+\.\d+)', re.DOTALL)

        for m in re.finditer(summary_regex, output):
            ws['G'][test_ids[id]].value = '{0:.5f}'.format(
                float(m.group('psnr_diff')))
            ws['H'][test_ids[id]].value = '{0:.5f}'.format(
                float(m.group('deviation')))

def run_test_cycle(id, transforms_str, repetitions, train_size, test_size):
    import prepare_data
    import sr_train
    import sr_test
    
    transform_args = {
        'Sc': '--scale',
        'Ro': '--rotate',
        'Tr': '--translate',
        'Fh': '--flip_horizontal',
        'Fv': '--flip_vertical',
        'La': '--tilt_angle',
        'Ba': '--additive_brightness',
        'Bm': '--brightness',
    }
    transform_regex = re.compile(
        r"(?P<t>\w+)\((?P<v>(-?\d+(\.\d+)?(,\s*-?\d+(\.\d+)?)*))?\)")
    models_regex = re.compile('model_epoch_(?P<epoch>\d+).pth')

    epochs_nonaugmented = 80
    epochs_augmented = 80

    os.makedirs('datasets/{}'.format(id), exist_ok=True)
    os.makedirs('tests/{}'.format(id), exist_ok=True)
    os.makedirs('checkpoints/{}'.format(id), exist_ok=True)
    
    for i in range(1, repetitions+1):
        args = ['--image_folder', opt.training_images,
                '--sample_size', str(train_size),
                '--save_images', '--save_dataset']
        if opt.train_csv:
            args.extend(['--dataset_csv', opt.train_csv])
        prepare_data.main(args)
        
        args = ['--image_folder', opt.training_images,
                '--dataset_csv', 'dataset.csv', '--save_images']
        transforms_list = [i.strip() for i in transforms_str.split(';')]
        for transform_combination in transforms_list:
            transforms = [i.strip() for i in transform_combination.split('&')]
            aug_args = args
            for transform in transforms:
                m = transform_regex.match(transform)
                if m:
                    aug_args = aug_args + [transform_args[m.group('t')]]
                    if m.group('v'):
                        if m.group('t') == 'Sc':
                            val = str(1 / float(m.group('v')))
                            aug_args.append(val)
                        elif m.group('t') == 'Bm' or m.group('t') == 'Tr':
                            vals = m.group('v').split(',')
                            if m.group('t') == 'Tr':
                                vals = [str(float(val)*2) for val in vals]
                            aug_args.extend(vals)
                        else:
                            aug_args.append(m.group('v'))
            prepare_data.main(aug_args)

        if str(train_size) not in saved_models:
            # Generate nonaugmented dataset consisting of original images
            # downsampled by 2
            args = ['--image_folder', 'generated', '--scale', '2',
                    '--hdf5_path', 'train.h5', '--dataset_csv', 'dataset.csv']
            prepare_data.main(args)
            
            # Start training for nonaugmented dataset
            args = ['--threads', '0', '--cuda',
                    '--nEpochs', str(epochs_nonaugmented)]
            # Use auto-termination for first repetition only
            if i == 1:
                args = args +  ['--test_images', opt.test_images,
                                '--sample_size', '100']
            sr_train.main(args)
            
            # Test model and save results
            if i == 1:
                epochs = [int(models_regex.match(f).group('epoch')) \
                          for f in os.listdir('checkpoint') \
                          if models_regex.match(f)]
                # Update the epoch limit once number of epochs required to
                # fully train the network is determined from the first repetition
                epochs_nonaugmented = max(epochs)
            args = ['--image_folder', opt.test_images,
                    '--sample_size', str(test_size),
                    '--model', 'checkpoint/model_epoch_{}.pth'.format(epochs_nonaugmented),
                    '--cuda', '--save_test', '--save_result']
            args = args + ['--scale'] + scale_factors
            if opt.test_csv:
                args.extend(['--load_test', opt.test_csv])
            sr_test.main(args)
        
            # Save HDF5 nonaugmented training dataset, models and results
            move('dataset.csv', 'datasets/{}/dataset_{}.csv'.format(id, i))
            move('train.h5',
                 'datasets/{}/train_nonaugmented_{}.h5'.format(id, i))
            move('results.csv',
                 'results/{}/results_nonaugmented_{}.csv'.format(id, i))
            move('checkpoint',
                 'checkpoints/{}/checkpoint_nonaugmented_{}'.format(id, i))
        else:
            dest = 'checkpoints/{}/checkpoint_nonaugmented_{}'.format(id, i)
            os.makedirs(dest, exist_ok=True)
            copy2(saved_models[str(train_size)][3], dest)
            copy2(saved_models[str(train_size)][0],
                  'datasets/{}/dataset_{}.csv'.format(id, i))
            copy2(saved_models[str(train_size)][1],
                  'datasets/{}/train_nonaugmented_{}.h5'.format(id, i))
            copy2(saved_models[str(train_size)][2],
                  'results/{}/results_nonaugmented_{}.csv'.format(id, i))            

        # Save tests
        test = 'tests/{}/test_{}.csv'.format(id, i)
        if str(test_size) not in saved_tests:
            move('test.csv', test)
        else:
            copy2(saved_tests[str(test_size)], test)
        
        # Keep a record of models to reuse based on training dataset size
        if opt.reuse_model and str(train_size) not in saved_models:
            saved_models[str(train_size)] = (
                'datasets/{}/dataset_{}.csv'.format(id, i),
                'datasets/{}/train_nonaugmented_{}.h5'.format(id, i),
                'results/{}/results_nonaugmented_{}.csv'.format(id, i),
                'checkpoints/{}/checkpoint_nonaugmented_{}/model_epoch_{}.pth'.format(id, i, epochs_nonaugmented))

        # Keep a record of tests to reuse based on test dataset size
        if opt.reuse_test and str(test_size) not in saved_tests:
            saved_tests[str(test_size)] = test

        # Generate augmented dataset consisting of original and
        # transformed images downsampled by 2
        args = ['--image_folder', 'generated', '--scale', '2',
                '--hdf5_path', 'train.h5']
        prepare_data.main(args)

        # Start training for augmented dataset
        args = ['--threads', '0', '--cuda',
                '--nEpochs', str(epochs_augmented)]
        # Use auto-termination for first repetition only
        if i == 1:
            args = args +  ['--test_images', opt.test_images,
                            '--sample_size', '100']
        sr_train.main(args)

        # Test model and save results
        if i == 1:
            epochs = [int(models_regex.match(f).group('epoch')) \
                      for f in os.listdir('checkpoint') \
                      if models_regex.match(f)]
            # Update the epoch limit once number of epochs required to
            # fully train the network is determined from the first repetition
            ws['F'][test_ids[id]].value = epochs_augmented = max(epochs)
        
        args = ['--image_folder', opt.test_images,
                '--sample_size', str(test_size),
                '--model', 'checkpoint/model_epoch_{}.pth'.format(epochs_augmented),
                '--cuda', '--load_test', test, '--save_result']
        args = args + ['--scale'] + scale_factors
        sr_test.main(args)

        # Save HDF5 augmented training dataset and results
        move('train.h5', 'datasets/{}/train_augmented_{}.h5'.format(id, i))
        move('results.csv', 'results/{}/results_augmented_{}.csv'.format(id, i))
        move('checkpoint',
             'checkpoints/{}/checkpoint_augmented_{}'.format(id, i))

        # Delete images for the next iteration
        rmtree('generated')

def main():
    transforms_list = []
    repetitions_list = []
    train_dataset_sizes = []
    test_dataset_sizes = []

    for row, col in enumerate(ws['A'][1:], 1):
        test_ids[col.value] = row
    for col in ws['B'][1:]:
        transforms_list.append(col.value)
    for col in ws['C'][1:]:
        repetitions_list.append(col.value)
    for col in ws['D'][1:]:
        train_dataset_sizes.append(col.value)
    for col in ws['E'][1:]:
        test_dataset_sizes.append(col.value)

    tests = OrderedDict(zip(test_ids.keys(),
                            zip(transforms_list, repetitions_list,
                                train_dataset_sizes, test_dataset_sizes)))

    with open("hdf5.json", "r") as config:
        global scale_factors
        scale_factors = json.load(config)["upscale_factors"]
        scale_factors = [str(scale) for scale in scale_factors]
    
    for id, params in tests.items():
        if opt.test_ids and str(id) not in opt.test_ids:
            continue
        
        os.makedirs('results/{}'.format(id), exist_ok=True)

        if not opt.analyze_only:
            run_test_cycle(id, params[0], params[1], params[2], params[3])
        
        # Collect results for the test
        analyze_results(id)

        wb.save(opt.worksheet)
    
if __name__ == "__main__":
    main()
